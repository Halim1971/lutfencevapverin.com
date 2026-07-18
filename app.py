import json
import os
import re
import secrets
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from urllib.parse import quote

from flask import (
    Flask,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
)
from itsdangerous import BadSignature, URLSafeTimedSerializer
from openpyxl import Workbook, load_workbook
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
UPLOAD_DIR = Path(os.environ.get("UPLOAD_PATH", BASE_DIR / "uploads"))
DB_PATH = Path(os.environ.get("DATABASE_PATH", INSTANCE_DIR / "lcv.sqlite3"))
ALLOWED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
ALLOWED_IMPORT_EXTENSIONS = {".json"}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-change-me")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("SESSION_COOKIE_SECURE", "0") == "1",
    PREFERRED_URL_SCHEME=os.environ.get("PREFERRED_URL_SCHEME", "http"),
)
if os.environ.get("TRUST_PROXY", "0") == "1":
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
API_TOKEN_SALT = os.environ.get("API_TOKEN_SALT", "lcv-api-token-v1")
REGISTRATION_ENABLED = False


def now_iso():
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def parse_dt(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


@app.after_request
def prevent_stale_dynamic_pages(response):
    if response.content_type and response.content_type.startswith("text/html"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


def init_db():
    INSTANCE_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(exist_ok=True)
    (UPLOAD_DIR / "invitations").mkdir(parents=True, exist_ok=True)
    db = get_db()
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            full_name TEXT,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'couple',
            is_beta_user INTEGER NOT NULL DEFAULT 0,
            beta_expires_at TEXT,
            lifetime_free INTEGER NOT NULL DEFAULT 0,
            is_access_blocked INTEGER NOT NULL DEFAULT 0,
            beta_login_password TEXT,
            login_password TEXT,
            managed_by_user_id INTEGER,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS invitations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL UNIQUE,
            image_path TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS guests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            full_name TEXT NOT NULL,
            phone TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'not_sent',
            last_sent_at TEXT,
            first_sent_at TEXT,
            responded_at TEXT,
            response_token TEXT NOT NULL UNIQUE,
            is_direct_participant INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS settings (
            user_id INTEGER PRIMARY KEY,
            no_response_after_hours INTEGER NOT NULL DEFAULT 48,
            invitation_message TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS guest_send_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            guest_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            sent_at TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(guest_id) REFERENCES guests(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS account_deletion_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            requested_at TEXT NOT NULL
        );
        """
    )
    ensure_column("users", "full_name", "TEXT")
    ensure_column("users", "is_beta_user", "INTEGER NOT NULL DEFAULT 0")
    ensure_column("users", "beta_expires_at", "TEXT")
    ensure_column("users", "lifetime_free", "INTEGER NOT NULL DEFAULT 0")
    ensure_column("users", "is_access_blocked", "INTEGER NOT NULL DEFAULT 0")
    ensure_column("users", "beta_login_password", "TEXT")
    ensure_column("users", "login_password", "TEXT")
    ensure_column("users", "managed_by_user_id", "INTEGER")
    ensure_column("users", "organization_company_name", "TEXT")
    ensure_column("settings", "invitation_message", "TEXT")
    ensure_column("guests", "is_direct_participant", "INTEGER NOT NULL DEFAULT 0")
    db.commit()


def ensure_column(table, column, definition):
    db = get_db()
    columns = [row["name"] for row in db.execute(f"PRAGMA table_info({table})").fetchall()]
    if column not in columns:
        db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


@app.cli.command("init-db")
def init_db_command():
    init_db()
    print("Veritabanı hazır.")


@app.cli.command("seed")
def seed_command():
    init_db()
    db = get_db()
    seed_user("admin@lutfencevapverin.com", "123456", "super_admin", "Süper Admin")
    seed_user("demo@lutfencevapverin.com", "123456", "couple", "Demo Kullanıcı")
    db.commit()
    print("Seed tamamlandı: admin@lutfencevapverin.com / 123456")


def seed_user(email, password, role, full_name=None):
    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
    if existing:
        return existing["id"]
    cur = db.execute(
        "INSERT INTO users (email, full_name, password_hash, role, created_at) VALUES (?, ?, ?, ?, ?)",
        (email, full_name, generate_password_hash(password), role, now_iso()),
    )
    user_id = cur.lastrowid
    db.execute(
        "INSERT OR IGNORE INTO settings (user_id, no_response_after_hours) VALUES (?, 48)",
        (user_id,),
    )
    db.execute(
        "INSERT OR IGNORE INTO invitations (user_id, created_at, updated_at) VALUES (?, ?, ?)",
        (user_id, now_iso(), now_iso()),
    )
    return user_id


def generate_beta_password():
    return f"{secrets.randbelow(10000):04d}"


def create_beta_user(email):
    email = str(email or "").strip().lower()
    if not email or "@" not in email or email.startswith("@") or email.endswith("@"):
        raise ValueError("Geçerli bir email adresi girin.")

    db = get_db()
    if db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone():
        raise LookupError("Bu email zaten kayıtlı.")

    password = generate_beta_password()
    user_id = seed_user(email, password, "couple")
    db.execute(
        """UPDATE users
           SET is_beta_user = 1, is_access_blocked = 0, beta_login_password = ?
           WHERE id = ?""",
        (password, user_id),
    )
    db.commit()
    return user_id, password


def create_managed_account(email, role="couple", managed_by_user_id=None):
    email = str(email or "").strip().lower()
    if not email or "@" not in email or email.startswith("@") or email.endswith("@"):
        raise ValueError("Geçerli bir email adresi girin.")
    if role not in {"couple", "organizer"}:
        raise ValueError("Geçersiz kullanıcı tipi.")
    db = get_db()
    if db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone():
        raise LookupError("Bu email zaten kayıtlı.")
    password = generate_beta_password()
    user_id = seed_user(email, password, role)
    db.execute(
        """UPDATE users SET login_password = ?, managed_by_user_id = ?,
           is_access_blocked = 0 WHERE id = ?""",
        (password, managed_by_user_id, user_id),
    )
    db.commit()
    return user_id, password


@app.before_request
def ensure_db():
    init_db()


def current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()


def user_payload(user):
    access = user_access_payload(user)
    return {
        "id": user["id"],
        "email": user["email"],
        "full_name": user["full_name"],
        "organization_company_name": user["organization_company_name"],
        "role": user["role"],
        "created_at": user["created_at"],
        "is_beta_user": bool(user["is_beta_user"]),
        "beta_expires_at": user["beta_expires_at"],
        "lifetime_free": bool(user["lifetime_free"]),
        "is_access_blocked": bool(user["is_access_blocked"]),
        "can_use_app": access["can_use_app"],
        "access_type": access["access_type"],
    }


def user_access_payload(user):
    """Merkezi uygulama erişim kararı; ödeme türleri ileride buraya eklenecek."""
    if user["role"] == "super_admin":
        return {"can_use_app": True, "access_type": "super_admin"}
    if bool(user["is_access_blocked"]):
        return {"can_use_app": False, "access_type": "blocked"}
    if bool(user["is_beta_user"]):
        if bool(user["is_access_blocked"]):
            return {"can_use_app": False, "access_type": "beta_blocked"}
        return {"can_use_app": True, "access_type": "beta"}
    return {"can_use_app": True, "access_type": "normal"}


def can_use_app(user):
    return user_access_payload(user)["can_use_app"]


def api_serializer():
    return URLSafeTimedSerializer(app.config["SECRET_KEY"], salt=API_TOKEN_SALT)


def make_api_token(user_id, *, admin_override=False, issued_by=None):
    payload = {"user_id": user_id}
    if admin_override:
        payload.update({"admin_override": True, "issued_by": issued_by})
    return api_serializer().dumps(payload)


def api_current_user():
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return None
    token = auth_header.removeprefix("Bearer ").strip()
    if not token:
        return None
    try:
        data = api_serializer().loads(token)
    except BadSignature:
        return None
    user_id = data.get("user_id")
    if not user_id:
        return None
    g.api_token_data = data
    return get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()


def api_login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = api_current_user()
        if not user:
            return jsonify({"error": "unauthorized"}), 401
        if not can_use_app(user) and not bool(getattr(g, "api_token_data", {}).get("admin_override")):
            return jsonify({"error": "Beta kullanıcı erişiminiz durdurulmuştur."}), 403
        g.api_user = user
        return fn(*args, **kwargs)

    return wrapper


def api_admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = api_current_user()
        if not user:
            return jsonify({"error": "unauthorized"}), 401
        if user["role"] != "super_admin":
            return jsonify({"error": "Bu işlem için yetkiniz yok."}), 403
        g.api_user = user
        return fn(*args, **kwargs)

    return wrapper


def api_manager_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = api_current_user()
        if not user:
            return jsonify({"error": "unauthorized"}), 401
        if user["role"] not in ("super_admin", "organizer"):
            return jsonify({"error": "Bu işlem için yetkiniz yok."}), 403
        if not can_use_app(user):
            return jsonify({"error": "Kullanıcı erişiminiz durdurulmuştur."}), 403
        g.api_user = user
        return fn(*args, **kwargs)

    return wrapper


def login_required(role=None):
    def decorator(fn):
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                return redirect(url_for("login"))
            if role and user["role"] != role:
                flash("Bu sayfaya erişim yetkiniz yok.", "error")
                return redirect(url_for("home"))
            if not can_use_app(user):
                session.clear()
                flash("Beta kullanıcı erişiminiz durdurulmuştur.", "error")
                return redirect(url_for("login"))
            profile_missing = (
                user["role"] == "couple" and not user["full_name"]
            ) or (
                user["role"] == "organizer" and not user["organization_company_name"]
            )
            if (
                profile_missing
                and request.endpoint != "complete_profile"
                and request.endpoint != "logout"
                and not request.endpoint.startswith("static")
            ):
                return redirect(url_for("complete_profile"))
            return fn(*args, **kwargs)

        wrapper.__name__ = fn.__name__
        return wrapper

    return decorator


def normalize_phone(raw):
    digits = re.sub(r"\D+", "", str(raw or ""))
    if not digits:
        return ""
    if digits.startswith("00"):
        digits = digits[2:]
    if digits.startswith("90"):
        return digits
    if digits.startswith("0") and len(digits) == 11:
        return "9" + digits
    if digits.startswith("5") and len(digits) == 10:
        return "90" + digits
    return digits


def display_status(guest):
    status = guest["status"]
    if status == "attending":
        return "Katılıyor"
    if status == "not_attending":
        return "Katılamıyor"
    if status == "cancelled_after_attending":
        return "Katılım onaylandı ama sonra iptal edildi"
    if status == "not_sent":
        return "Gönderilmedi"
    if is_phone_call_needed(guest):
        return "Telefonla aranmalı"
    if status == "followup_waiting":
        return "Cevap Bekleniyor"
    if is_no_response(guest):
        return "Cevap Vermedi"
    return "Cevap Bekleniyor"


def get_setting(user_id):
    row = get_db().execute(
        "SELECT no_response_after_hours FROM settings WHERE user_id = ?", (user_id,)
    ).fetchone()
    if row:
        return int(row["no_response_after_hours"])
    get_db().execute(
        "INSERT OR IGNORE INTO settings (user_id, no_response_after_hours) VALUES (?, 48)",
        (user_id,),
    )
    get_db().commit()
    return 48


def get_invitation_message(user_id):
    row = get_db().execute(
        "SELECT invitation_message FROM settings WHERE user_id = ?", (user_id,)
    ).fetchone()
    if row and row["invitation_message"]:
        return row["invitation_message"]
    return "Düğünümüze katılım durumunuzu bildirmek için lütfen aşağıdaki linkten cevap verin: {link}"


def is_no_response(guest):
    if not guest["last_sent_at"]:
        return False
    if guest["status"] == "followup_waiting":
        return not is_phone_call_needed(guest)
    if guest["status"] != "sent_waiting":
        return False
    sent_at = parse_dt(guest["last_sent_at"])
    if not sent_at:
        return False
    hours = get_setting(guest["user_id"])
    return datetime.now() >= sent_at + timedelta(hours=hours)


def is_phone_call_needed(guest):
    if guest["status"] != "followup_waiting" or not guest["last_sent_at"]:
        return False
    sent_at = parse_dt(guest["last_sent_at"])
    if not sent_at:
        return False
    hours = get_setting(guest["user_id"])
    return datetime.now() >= sent_at + timedelta(hours=hours)


def guest_counts(user_id):
    guests = get_db().execute("SELECT * FROM guests WHERE user_id = ?", (user_id,)).fetchall()
    invitation_guests = [g for g in guests if not g["is_direct_participant"]]
    return {
        "total": len(invitation_guests),
        "attending": sum(1 for g in guests if g["status"] == "attending"),
        "not_attending": sum(1 for g in invitation_guests if g["status"] == "not_attending"),
        "no_response": sum(1 for g in invitation_guests if is_no_response(g)),
        "phone_followup": sum(1 for g in invitation_guests if is_phone_call_needed(g)),
        "waiting": sum(1 for g in invitation_guests if g["status"] == "sent_waiting" and not is_no_response(g)),
        "not_sent": sum(1 for g in invitation_guests if g["status"] == "not_sent"),
    }


def invitation_for(user_id):
    row = get_db().execute("SELECT * FROM invitations WHERE user_id = ?", (user_id,)).fetchone()
    if row:
        return row
    get_db().execute(
        "INSERT INTO invitations (user_id, created_at, updated_at) VALUES (?, ?, ?)",
        (user_id, now_iso(), now_iso()),
    )
    get_db().commit()
    return get_db().execute("SELECT * FROM invitations WHERE user_id = ?", (user_id,)).fetchone()


def invitation_url(invitation):
    if not invitation or not invitation["image_path"]:
        return None
    return url_for("uploaded_file", filename=invitation["image_path"])


def add_guest(user_id, full_name, phone):
    phone = normalize_phone(phone)
    full_name = " ".join(str(full_name or "").split())
    if not full_name or not phone:
        return False, "İsim ve telefon zorunlu."
    db = get_db()
    existing = db.execute(
        "SELECT id FROM guests WHERE user_id = ? AND phone = ?", (user_id, phone)
    ).fetchone()
    if existing:
        return False, "Bu telefon numarası zaten listede var."
    db.execute(
        """
        INSERT INTO guests
        (user_id, full_name, phone, status, response_token, is_direct_participant, created_at, updated_at)
        VALUES (?, ?, ?, 'not_sent', ?, 0, ?, ?)
        """,
        (user_id, full_name, phone, secrets.token_urlsafe(24), now_iso(), now_iso()),
    )
    return True, "Davetli eklendi."


def add_direct_participant(user_id, full_name, phone):
    phone = normalize_phone(phone)
    full_name = " ".join(str(full_name or "").split())
    if not full_name or not phone:
        return False, "İsim ve telefon zorunlu."
    db = get_db()
    existing = db.execute(
        "SELECT id FROM guests WHERE user_id = ? AND phone = ?", (user_id, phone)
    ).fetchone()
    if existing:
        return False, "Bu telefon numarası zaten listede var."
    db.execute(
        """
        INSERT INTO guests
        (user_id, full_name, phone, status, responded_at, response_token, is_direct_participant, created_at, updated_at)
        VALUES (?, ?, ?, 'attending', ?, ?, 1, ?, ?)
        """,
        (user_id, full_name, phone, now_iso(), secrets.token_urlsafe(24), now_iso(), now_iso()),
    )
    return True, "Katılımcı eklendi."


def personal_link(guest, external=False):
    return url_for("respond", token=guest["response_token"], _external=external)


def whatsapp_link(guest):
    link = personal_link(guest, external=True)
    template = get_invitation_message(guest["user_id"])
    text = template.replace("{link}", link)
    if link not in text:
        text = f"{text}\n{link}"
    return f"https://wa.me/{guest['phone']}?text={quote(text)}"


def record_send_log(guest_id, user_id, sent_at):
    get_db().execute(
        """
        INSERT INTO guest_send_logs (guest_id, user_id, sent_at, created_at)
        VALUES (?, ?, ?, ?)
        """,
        (guest_id, user_id, sent_at, now_iso()),
    )


def send_history(guest, limit=3):
    rows = get_db().execute(
        """
        SELECT sent_at FROM guest_send_logs
        WHERE guest_id = ? AND user_id = ?
        ORDER BY sent_at ASC, id ASC
        """,
        (guest["id"], guest["user_id"]),
    ).fetchall()
    history = [row["sent_at"] for row in rows]
    if not history:
        for value in (guest["first_sent_at"], guest["last_sent_at"]):
            if value and value not in history:
                history.append(value)
    return history[-limit:]


def load_guests(user_id, filter_name=None):
    rows = get_db().execute(
        "SELECT * FROM guests WHERE user_id = ? ORDER BY created_at DESC, id DESC", (user_id,)
    ).fetchall()
    if filter_name == "attending":
        return [g for g in rows if g["status"] == "attending"]
    if filter_name == "not_attending":
        return [g for g in rows if g["status"] == "not_attending"]
    if filter_name == "no_response":
        return [g for g in rows if is_no_response(g)]
    if filter_name == "phone_followup":
        return [g for g in rows if is_phone_call_needed(g)]
    if filter_name == "waiting":
        return [g for g in rows if g["status"] == "sent_waiting" and not is_no_response(g)]
    if filter_name == "not_sent":
        return [g for g in rows if g["status"] == "not_sent" and not g["is_direct_participant"]]
    return [g for g in rows if not g["is_direct_participant"]]


@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    data = request.get_json(silent=True) or {}
    email = str(data.get("email", "")).strip().lower()
    password = str(data.get("password", ""))
    user = get_db().execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "Email veya şifre hatalı."}), 401
    if not can_use_app(user):
        return jsonify({"error": "Beta kullanıcı erişiminiz durdurulmuştur."}), 403
    return jsonify({"token": make_api_token(user["id"]), "user": user_payload(user)})


@app.route("/api/auth/register", methods=["POST"])
def api_auth_register():
    if not REGISTRATION_ENABLED:
        return jsonify({"error": "Yeni kayıtlar geçici olarak kapalıdır."}), 403
    data = request.get_json(silent=True) or {}
    email = str(data.get("email", "")).strip().lower()
    password = str(data.get("password", ""))
    if not email or len(password) < 6:
        return jsonify({"error": "Email ve en az 6 karakter şifre girin."}), 400
    try:
        user_id = seed_user(email, password, "couple")
        get_db().commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "Bu email zaten kayıtlı."}), 409
    user = get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return jsonify({"token": make_api_token(user["id"]), "user": user_payload(user)}), 201


@app.route("/api/me", methods=["GET", "POST"])
@api_login_required
def api_me():
    if request.method == "POST":
        full_name = " ".join(str((request.get_json(silent=True) or {}).get("full_name", "")).split())
        if len(full_name) < 3:
            message = "Lütfen organizasyon şirketi adını girin." if g.api_user["role"] == "organizer" else "Lütfen isim soyisim girin."
            return jsonify({"error": message}), 400
        if g.api_user["role"] == "organizer":
            get_db().execute("UPDATE users SET organization_company_name = ? WHERE id = ?", (full_name, g.api_user["id"]))
        else:
            get_db().execute("UPDATE users SET full_name = ? WHERE id = ?", (full_name, g.api_user["id"]))
        get_db().commit()
        g.api_user = get_db().execute("SELECT * FROM users WHERE id = ?", (g.api_user["id"],)).fetchone()
    return jsonify({"user": user_payload(g.api_user)})


@app.route("/api/dashboard")
@api_login_required
def api_dashboard():
    user = g.api_user
    if user["role"] == "super_admin":
        return jsonify(
            {
                "user": user_payload(user),
                "admin": True,
                "counts": {
                    "total": 0,
                    "attending": 0,
                    "not_attending": 0,
                    "no_response": 0,
                    "phone_followup": 0,
                    "waiting": 0,
                    "not_sent": 0,
                },
                "invitation": {"image_url": None, "message": None},
            }
        )
    inv = invitation_for(user["id"])
    return jsonify(
        {
            "user": user_payload(user),
            "counts": guest_counts(user["id"]),
            "invitation": {
                "image_url": invitation_url(inv),
                "message": get_invitation_message(user["id"]),
            },
        }
    )


def api_guest_payload(guest):
    return {
        "id": guest["id"],
        "full_name": guest["full_name"],
        "phone": guest["phone"],
        "status": guest["status"],
        "status_label": display_status(guest),
        "is_direct_participant": bool(guest["is_direct_participant"]),
        "responded_at": guest["responded_at"],
        "send_history": send_history(guest),
        "can_send": guest["status"] not in ("attending", "not_attending", "cancelled_after_attending") and not is_phone_call_needed(guest),
        "send_label": "Tekrar Gönder" if guest["last_sent_at"] else "Gönder",
    }


@app.route("/api/guests", methods=["GET", "POST"])
@api_login_required
def api_guests():
    user = g.api_user
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        ok, message = add_guest(user["id"], data.get("full_name"), data.get("phone"))
        if not ok:
            return jsonify({"error": message}), 400
        get_db().commit()
    filter_name = request.args.get("filter")
    guests = load_guests(user["id"], filter_name)
    return jsonify({"guests": [api_guest_payload(guest) for guest in guests]})


@app.route("/api/guests/batch", methods=["POST"])
@api_login_required
def api_guests_batch():
    data = request.get_json(silent=True) or {}
    contacts = data.get("contacts", [])
    if not isinstance(contacts, list) or not contacts:
        return jsonify({"error": "Lütfen en az bir kişi seçin."}), 400
    added = skipped = 0
    for contact in contacts:
        if not isinstance(contact, dict):
            skipped += 1
            continue
        ok, _ = add_guest(
            g.api_user["id"], contact.get("full_name"), contact.get("phone")
        )
        added += int(ok)
        skipped += int(not ok)
    get_db().commit()
    return jsonify({
        "added": added,
        "skipped": skipped,
        "message": f"Rehber aktarımı tamamlandı. Eklenen: {added}, atlanan: {skipped}",
    })


@app.route("/api/invitation", methods=["GET", "POST"])
@api_login_required
def api_invitation():
    user = g.api_user
    if request.method == "POST":
        if request.is_json:
            message = str((request.get_json(silent=True) or {}).get("message", "")).strip()
            if not message:
                return jsonify({"error": "Davetiye mesajı boş olamaz."}), 400
            get_db().execute(
                """INSERT INTO settings (user_id, no_response_after_hours, invitation_message)
                VALUES (?, ?, ?) ON CONFLICT(user_id) DO UPDATE SET invitation_message = excluded.invitation_message""",
                (user["id"], get_setting(user["id"]), message),
            )
            get_db().commit()
        else:
            file = request.files.get("invitation")
            if not file or not file.filename:
                return jsonify({"error": "Lütfen bir görsel seçin."}), 400
            ext = Path(file.filename).suffix.lower()
            if ext not in ALLOWED_IMAGE_EXTENSIONS:
                return jsonify({"error": "Lütfen PNG, JPG veya WEBP formatında davetiye yükleyin."}), 400
            filename = secure_filename(f"user_{user['id']}_{secrets.token_hex(8)}{ext}")
            rel_path = f"invitations/{filename}"
            file.save(UPLOAD_DIR / rel_path)
            get_db().execute(
                """INSERT INTO invitations (user_id, image_path, created_at, updated_at)
                VALUES (?, ?, ?, ?) ON CONFLICT(user_id) DO UPDATE SET image_path = excluded.image_path, updated_at = excluded.updated_at""",
                (user["id"], rel_path, now_iso(), now_iso()),
            )
            get_db().commit()
    invitation = invitation_for(user["id"])
    return jsonify({
        "image_url": invitation_url(invitation),
        "message": get_invitation_message(user["id"]),
    })


@app.route("/api/guests/excel", methods=["POST"])
@api_login_required
def api_guests_excel():
    file = request.files.get("excel_file")
    if not file or not file.filename or Path(file.filename).suffix.lower() != ".xlsx":
        return jsonify({"error": "Lütfen XLSX formatında Excel dosyası seçin."}), 400
    added = skipped = 0
    try:
        worksheet = load_workbook(file, read_only=True, data_only=True).active
        for row in worksheet.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2 or not row[0] or not row[1]:
                continue
            ok, _ = add_guest(g.api_user["id"], row[0], row[1])
            added += int(ok)
            skipped += int(not ok)
    except Exception:
        return jsonify({"error": "Excel dosyası okunamadı."}), 400
    get_db().commit()
    return jsonify({"message": f"Excel yükleme tamamlandı. Eklenen: {added}, atlanan: {skipped}"})


@app.route("/api/guests/actions", methods=["POST"])
@api_login_required
def api_guest_actions():
    data = request.get_json(silent=True) or {}
    ids = [int(value) for value in data.get("guest_ids", []) if str(value).isdigit()]
    action = data.get("action")
    if not ids:
        return jsonify({"error": "Lütfen en az bir davetli seçin."}), 400
    placeholders = ",".join("?" for _ in ids)
    rows = get_db().execute(
        f"SELECT * FROM guests WHERE user_id = ? AND id IN ({placeholders})", [g.api_user["id"], *ids]
    ).fetchall()
    if action == "delete":
        get_db().execute(f"DELETE FROM guest_send_logs WHERE user_id = ? AND guest_id IN ({placeholders})", [g.api_user["id"], *ids])
        get_db().execute(f"DELETE FROM guests WHERE user_id = ? AND id IN ({placeholders})", [g.api_user["id"], *ids])
        get_db().commit()
        return jsonify({"message": f"Seçili {len(rows)} davetli listeden çıkarıldı."})
    if action == "cancel_attending":
        get_db().execute(
            f"UPDATE guests SET status = 'cancelled_after_attending', updated_at = ? WHERE user_id = ? AND id IN ({placeholders})",
            [now_iso(), g.api_user["id"], *ids],
        )
        get_db().commit()
        return jsonify({"message": "Seçili katılımcılar listeden çıkarıldı."})
    if action == "send":
        links = []
        for guest in rows:
            sent_at = now_iso()
            first_sent = guest["first_sent_at"] or sent_at
            get_db().execute(
                """UPDATE guests SET status = CASE WHEN status IN ('attending','not_attending') THEN status ELSE 'sent_waiting' END,
                first_sent_at = ?, last_sent_at = ?, updated_at = ? WHERE id = ?""",
                (first_sent, sent_at, sent_at, guest["id"]),
            )
            record_send_log(guest["id"], g.api_user["id"], sent_at)
            links.append({"id": guest["id"], "full_name": guest["full_name"], "url": whatsapp_link(guest)})
        get_db().commit()
        return jsonify({"links": links})
    return jsonify({"error": "Geçersiz işlem."}), 400


@app.route("/api/guests/whatsapp-links", methods=["POST"])
@api_login_required
def api_guest_whatsapp_links():
    data = request.get_json(silent=True) or {}
    ids = [int(value) for value in data.get("guest_ids", []) if str(value).isdigit()]
    if not ids:
        return jsonify({"error": "Lütfen en az bir davetli seçin."}), 400
    placeholders = ",".join("?" for _ in ids)
    rows = get_db().execute(
        f"SELECT * FROM guests WHERE user_id = ? AND id IN ({placeholders}) ORDER BY created_at DESC, id DESC",
        [g.api_user["id"], *ids],
    ).fetchall()
    return jsonify({"links": [
        {"id": guest["id"], "full_name": guest["full_name"], "url": whatsapp_link(guest)}
        for guest in rows
    ]})


@app.route("/api/guests/<int:guest_id>/mark-sent", methods=["POST"])
@api_login_required
def api_mark_sent(guest_id):
    guest = get_db().execute(
        "SELECT * FROM guests WHERE id = ? AND user_id = ?", (guest_id, g.api_user["id"])
    ).fetchone()
    if not guest:
        return jsonify({"error": "Davetli bulunamadı."}), 404
    sent_at = now_iso()
    first_sent = guest["first_sent_at"] or sent_at
    next_status = "followup_waiting" if is_no_response(guest) else "sent_waiting"
    get_db().execute(
        """UPDATE guests SET status = CASE WHEN status IN ('attending','not_attending') THEN status ELSE ? END,
        first_sent_at = ?, last_sent_at = ?, updated_at = ? WHERE id = ?""",
        (next_status, first_sent, sent_at, sent_at, guest_id),
    )
    record_send_log(guest_id, g.api_user["id"], sent_at)
    get_db().commit()
    return jsonify({"message": "Gönderim kaydedildi."})


@app.route("/api/guests/<int:guest_id>/manual-attending", methods=["POST"])
@api_login_required
def api_manual_attending(guest_id):
    cursor = get_db().execute(
        "UPDATE guests SET status = 'attending', responded_at = ?, updated_at = ? WHERE id = ? AND user_id = ?",
        (now_iso(), now_iso(), guest_id, g.api_user["id"]),
    )
    if not cursor.rowcount:
        return jsonify({"error": "Davetli bulunamadı."}), 404
    get_db().commit()
    return jsonify({"message": "Davetli katılımı manuel onaylandı."})


@app.route("/api/attending", methods=["POST"])
@api_login_required
def api_add_attending():
    data = request.get_json(silent=True) or {}
    ok, message = add_direct_participant(g.api_user["id"], data.get("full_name"), data.get("phone"))
    get_db().commit()
    return jsonify({"message" if ok else "error": message}), 201 if ok else 400


@app.route("/api/settings", methods=["GET", "POST"])
@api_login_required
def api_settings():
    if request.method == "POST":
        try:
            hours = max(1, int((request.get_json(silent=True) or {}).get("hours")))
        except (TypeError, ValueError):
            return jsonify({"error": "Lütfen geçerli bir saat girin."}), 400
        get_db().execute(
            """INSERT INTO settings (user_id, no_response_after_hours) VALUES (?, ?)
            ON CONFLICT(user_id) DO UPDATE SET no_response_after_hours = excluded.no_response_after_hours""",
            (g.api_user["id"], hours),
        )
        get_db().commit()
    return jsonify({"hours": get_setting(g.api_user["id"])})


@app.route("/api/guests/import-json", methods=["POST"])
@api_login_required
def api_import_json():
    file = request.files.get("guest_file")
    if not file or not file.filename or Path(file.filename).suffix.lower() != ".json":
        return jsonify({"error": "Lütfen JSON formatında davetli dosyası seçin."}), 400
    try:
        payload = json.load(file)
        items = payload.get("guests", [])
    except (ValueError, AttributeError):
        return jsonify({"error": "JSON dosyası okunamadı."}), 400
    added = skipped = 0
    for item in items:
        ok, _ = add_guest(g.api_user["id"], item.get("full_name"), item.get("phone"))
        added += int(ok); skipped += int(not ok)
    get_db().commit()
    return jsonify({"message": f"İçe aktarma tamamlandı. Eklenen: {added}, atlanan: {skipped}"})


@app.route("/api/guests/export-json")
@api_login_required
def api_export_json():
    guests = load_guests(g.api_user["id"])
    return jsonify({"guests": [{"full_name": row["full_name"], "phone": row["phone"]} for row in guests]})


@app.route("/api/final-participants")
@api_login_required
def api_final_participants():
    rows = final_participant_rows(g.api_user["id"])
    participants = [
        {"full_name": row["full_name"], "phone": row["phone"], "responded_at": row["responded_at"]}
        for row in rows
    ]
    return jsonify({"participants": participants, "count": len(participants)})


def management_account_payload(user):
    return {
        "user": user_payload(user),
        "login_password": user["login_password"],
        "managed_by_user_id": user["managed_by_user_id"],
        "counts": guest_counts(user["id"]) if user["role"] == "couple" else None,
    }


@app.route("/api/management/accounts", methods=["GET", "POST"])
@api_manager_required
def api_management_accounts():
    viewer = g.api_user
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        role = str(data.get("role", "couple"))
        if viewer["role"] == "organizer":
            role = "couple"
            manager_id = viewer["id"]
        else:
            if role not in ("organizer", "couple"):
                return jsonify({"error": "Geçersiz kullanıcı türü."}), 400
            manager_id = None
        try:
            user_id, password = create_managed_account(data.get("email"), role, manager_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 409
        user = get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        return jsonify({
            "account": management_account_payload(user),
            "temporary_password": password,
        }), 201

    requested_role = request.args.get("role", "couple")
    managed_by = request.args.get("managed_by")
    if viewer["role"] == "organizer":
        rows = get_db().execute(
            "SELECT * FROM users WHERE role = 'couple' AND managed_by_user_id = ? ORDER BY created_at DESC, id DESC",
            (viewer["id"],),
        ).fetchall()
    elif requested_role == "organizer":
        rows = get_db().execute(
            "SELECT * FROM users WHERE role = 'organizer' ORDER BY created_at DESC, id DESC"
        ).fetchall()
    elif managed_by and str(managed_by).isdigit():
        rows = get_db().execute(
            "SELECT * FROM users WHERE role = 'couple' AND managed_by_user_id = ? ORDER BY created_at DESC, id DESC",
            (int(managed_by),),
        ).fetchall()
    else:
        rows = get_db().execute(
            "SELECT * FROM users WHERE role = 'couple' ORDER BY created_at DESC, id DESC"
        ).fetchall()
    return jsonify({"accounts": [management_account_payload(user) for user in rows]})


@app.route("/api/management/accounts/<int:user_id>/access", methods=["POST"])
@api_admin_required
def api_management_account_access(user_id):
    user = get_db().execute(
        "SELECT * FROM users WHERE id = ? AND role IN ('couple', 'organizer')", (user_id,)
    ).fetchone()
    if not user:
        return jsonify({"error": "Kullanıcı bulunamadı."}), 404
    blocked = bool((request.get_json(silent=True) or {}).get("blocked"))
    get_db().execute("UPDATE users SET is_access_blocked = ? WHERE id = ?", (int(blocked), user_id))
    get_db().commit()
    return jsonify({"message": "Kullanım engellendi." if blocked else "Kullanım açıldı."})


@app.route("/api/management/accounts/<int:user_id>", methods=["DELETE"])
@api_admin_required
def api_management_account_delete(user_id):
    if user_id == g.api_user["id"]:
        return jsonify({"error": "Kendi Süper Admin hesabınızı silemezsiniz."}), 400
    user = delete_managed_account_records(user_id)
    if not user:
        return jsonify({"error": "Kullanıcı bulunamadı."}), 404
    return jsonify({"message": "Kullanıcı ve bağlı kayıtları kalıcı olarak silindi."})


def delete_managed_account_records(user_id):
    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE id = ? AND role IN ('couple', 'organizer')", (user_id,)
    ).fetchone()
    if not user:
        return None

    invitation_path = None
    if user["role"] == "couple":
        invitation = db.execute(
            "SELECT image_path FROM invitations WHERE user_id = ?", (user_id,)
        ).fetchone()
        invitation_path = invitation["image_path"] if invitation else None
        db.execute("DELETE FROM guest_send_logs WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM guests WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM invitations WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM settings WHERE user_id = ?", (user_id,))
    else:
        db.execute(
            "UPDATE users SET managed_by_user_id = NULL WHERE managed_by_user_id = ?",
            (user_id,),
        )
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()

    if invitation_path:
        try:
            (UPLOAD_DIR / invitation_path).unlink(missing_ok=True)
        except OSError:
            app.logger.warning("Silinen kullanıcıya ait davetiye görseli kaldırılamadı: %s", invitation_path)
    return user


@app.route("/api/management/users/<int:user_id>/readonly")
@api_manager_required
def api_management_user_readonly(user_id):
    viewer = g.api_user
    user = get_db().execute(
        "SELECT * FROM users WHERE id = ? AND role = 'couple'", (user_id,)
    ).fetchone()
    allowed = user and (
        viewer["role"] == "super_admin"
        or user["managed_by_user_id"] == viewer["id"]
    )
    if not allowed:
        return jsonify({"error": "Bu kullanıcıyı görüntüleme yetkiniz yok."}), 403
    invitation = invitation_for(user_id)
    return jsonify({
        "user": user_payload(user),
        "counts": guest_counts(user_id),
        "guests": [api_guest_payload(guest) for guest in load_guests(user_id)],
        "invitation": {
            "image_url": invitation_url(invitation),
            "message": get_invitation_message(user_id),
        },
        "no_response_after_hours": get_setting(user_id),
        "final_participants": [api_guest_payload(guest) for guest in final_participant_rows(user_id)],
    })


@app.route("/api/admin/users")
@api_admin_required
def api_admin_users():
    user_type = request.args.get("type", "beta")
    want_beta = user_type == "beta"
    rows = get_db().execute(
        "SELECT * FROM users WHERE role = 'couple' AND is_beta_user = ? ORDER BY created_at DESC, id DESC",
        (1 if want_beta else 0,),
    ).fetchall()
    return jsonify({"users": [
        {"user": user_payload(user), "counts": guest_counts(user["id"])}
        for user in rows
    ]})


@app.route("/api/admin/users/<int:user_id>")
@api_admin_required
def api_admin_user_detail(user_id):
    user = get_db().execute(
        "SELECT * FROM users WHERE id = ? AND role = 'couple'", (user_id,)
    ).fetchone()
    if not user:
        return jsonify({"error": "Kullanıcı bulunamadı."}), 404
    invitation = invitation_for(user_id)
    return jsonify({
        "user": user_payload(user),
        "counts": guest_counts(user_id),
        "guests": [api_guest_payload(guest) for guest in load_guests(user_id)],
        "invitation": {"image_url": invitation_url(invitation)},
    })


@app.route("/api/admin/beta-users", methods=["POST"])
@api_admin_required
def api_admin_create_beta():
    email = str((request.get_json(silent=True) or {}).get("email", "")).strip().lower()
    if not email or "@" not in email:
        return jsonify({"error": "Geçerli bir email adresi girin."}), 400
    if get_db().execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone():
        return jsonify({"error": "Bu email zaten kayıtlı."}), 409
    password = str(secrets.randbelow(9000) + 1000)
    user_id = seed_user(email, password, "couple")
    get_db().execute(
        "UPDATE users SET is_beta_user = 1, is_access_blocked = 0 WHERE id = ?",
        (user_id,),
    )
    get_db().commit()
    user = get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return jsonify({"user": user_payload(user), "temporary_password": password}), 201


@app.route("/api/admin/users/<int:user_id>/access", methods=["POST"])
@api_admin_required
def api_admin_user_access(user_id):
    user = get_db().execute(
        "SELECT * FROM users WHERE id = ? AND role = 'couple'", (user_id,)
    ).fetchone()
    if not user:
        return jsonify({"error": "Kullanıcı bulunamadı."}), 404
    if not user["is_beta_user"]:
        return jsonify({"error": "Yalnızca beta kullanıcıların erişimi değiştirilebilir."}), 400
    blocked = bool((request.get_json(silent=True) or {}).get("blocked"))
    get_db().execute(
        "UPDATE users SET is_access_blocked = ? WHERE id = ?", (int(blocked), user_id)
    )
    get_db().commit()
    return jsonify({"message": "Kullanım deaktive edildi." if blocked else "Kullanım tekrar aktive edildi."})


@app.route("/api/admin/users/<int:user_id>/impersonate", methods=["POST"])
@api_admin_required
def api_admin_impersonate(user_id):
    user = get_db().execute(
        "SELECT * FROM users WHERE id = ? AND role = 'couple'", (user_id,)
    ).fetchone()
    if not user:
        return jsonify({"error": "Kullanıcı bulunamadı."}), 404
    return jsonify({
        "token": make_api_token(
            user_id, admin_override=True, issued_by=g.api_user["id"]
        ),
        "user": user_payload(user),
    })


@app.route("/")
def root():
    return redirect(url_for("home") if current_user() else url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = get_db().execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            if not can_use_app(user):
                flash("Beta kullanıcı erişiminiz durdurulmuştur.", "error")
                return render_template("login.html")
            session.clear()
            session["user_id"] = user["id"]
            if user["role"] == "super_admin":
                return redirect(url_for("admin"))
            if user["role"] == "organizer":
                if not user["organization_company_name"]:
                    return redirect(url_for("complete_profile"))
                return redirect(url_for("organizer_dashboard"))
            if not user["full_name"]:
                return redirect(url_for("complete_profile"))
            return redirect(url_for("home"))
        flash("Email veya şifre hatalı.", "error")
    return render_template("login.html")


@app.route("/complete-profile", methods=["GET", "POST"])
@login_required()
def complete_profile():
    user = current_user()
    if user["role"] == "super_admin":
        return redirect(url_for("admin"))
    if request.method == "POST":
        full_name = " ".join(request.form.get("full_name", "").split())
        if len(full_name) < 3:
            flash("Lütfen organizasyon şirketi adını girin." if user["role"] == "organizer" else "Lütfen isim soyisim girin.", "error")
            return render_template("complete_profile.html", user=user)
        if user["role"] == "organizer":
            get_db().execute("UPDATE users SET organization_company_name = ? WHERE id = ?", (full_name, user["id"]))
        else:
            get_db().execute("UPDATE users SET full_name = ? WHERE id = ?", (full_name, user["id"]))
        get_db().commit()
        flash("Bilgileriniz kaydedildi.", "success")
        return redirect(url_for("organizer_dashboard" if user["role"] == "organizer" else "home"))
    return render_template("complete_profile.html", user=user)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if not REGISTRATION_ENABLED:
        flash("Yeni kayıtlar geçici olarak kapalıdır.", "error")
        return redirect(url_for("login"))
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        if not email or len(password) < 6:
            flash("Email ve en az 6 karakter şifre girin.", "error")
            return render_template("register.html")
        try:
            seed_user(email, password, "couple")
            get_db().commit()
        except sqlite3.IntegrityError:
            flash("Bu email zaten kayıtlı.", "error")
            return render_template("register.html")
        flash("Hesap oluşturuldu. Giriş yapabilirsiniz.", "success")
        return redirect(url_for("login"))
    return render_template("register.html")


@app.route("/home")
@login_required()
def home():
    user = current_user()
    if user["role"] == "super_admin":
        return redirect(url_for("admin"))
    if user["role"] == "organizer":
        return redirect(url_for("organizer_dashboard"))
    return render_template("home.html", user=user, counts=guest_counts(user["id"]))


def owner_context(user_id):
    user = get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user or user["role"] != "couple":
        return None
    return {
        "user": user,
        "counts": guest_counts(user_id),
        "all_guests": load_guests(user_id),
        "attending": load_guests(user_id, "attending"),
        "not_attending": load_guests(user_id, "not_attending"),
        "no_response": load_guests(user_id, "no_response"),
        "invitation": invitation_for(user_id),
    }


@app.route("/invitation", methods=["GET", "POST"])
@login_required()
def invitation():
    user = current_user()
    if request.method == "POST":
        if "invitation_message" in request.form:
            message = request.form.get("invitation_message", "").strip()
            if not message:
                flash("Davetiye mesajı boş olamaz.", "error")
                return redirect(url_for("invitation"))
            get_db().execute(
                """
                INSERT INTO settings (user_id, no_response_after_hours, invitation_message)
                VALUES (?, ?, ?)
                ON CONFLICT(user_id) DO UPDATE SET invitation_message = excluded.invitation_message
                """,
                (user["id"], get_setting(user["id"]), message),
            )
            get_db().commit()
            flash("Davetiye mesajı kaydedildi.", "success")
            return redirect(url_for("invitation"))
        file = request.files.get("invitation")
        if not file or not file.filename:
            flash("Lütfen bir görsel seçin.", "error")
            return redirect(url_for("invitation"))
        ext = Path(file.filename).suffix.lower()
        if ext not in ALLOWED_IMAGE_EXTENSIONS:
            flash("Lütfen PNG, JPG veya WEBP formatında davetiye yükleyin.", "error")
            return redirect(url_for("invitation"))
        filename = secure_filename(f"user_{user['id']}_{secrets.token_hex(8)}{ext}")
        rel_path = f"invitations/{filename}"
        file.save(UPLOAD_DIR / rel_path)
        db = get_db()
        db.execute(
            """
            INSERT INTO invitations (user_id, image_path, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET image_path = excluded.image_path, updated_at = excluded.updated_at
            """,
            (user["id"], rel_path, now_iso(), now_iso()),
        )
        db.commit()
        flash("Davetiye yüklendi.", "success")
        return redirect(url_for("invitation"))
    inv = invitation_for(user["id"])
    return render_template(
        "invitation.html",
        invitation=inv,
        invitation_url=invitation_url(inv),
        invitation_message=get_invitation_message(user["id"]),
    )


@app.route("/guests", methods=["GET", "POST"])
@login_required()
def guests():
    user = current_user()
    if request.method == "POST":
        ok, message = add_guest(
            user["id"], request.form.get("full_name"), request.form.get("phone")
        )
        get_db().commit()
        flash(message, "success" if ok else "error")
        return redirect(url_for("guests"))
    return render_template(
        "guests.html",
        title="Davetli Ekle / Çıkar",
        guests=load_guests(user["id"]),
        counts=guest_counts(user["id"]),
        show_add=True,
    )


@app.route("/guests/<int:guest_id>/mark-sent", methods=["POST"])
@login_required()
def mark_sent(guest_id):
    user = current_user()
    guest = get_db().execute(
        "SELECT * FROM guests WHERE id = ? AND user_id = ?", (guest_id, user["id"])
    ).fetchone()
    if not guest:
        flash("Davetli bulunamadı.", "error")
        return redirect(url_for("guests"))
    sent_at = now_iso()
    first_sent = guest["first_sent_at"] or sent_at
    next_status = "followup_waiting" if is_no_response(guest) else "sent_waiting"
    get_db().execute(
        """
        UPDATE guests
        SET status = CASE WHEN status IN ('attending', 'not_attending') THEN status ELSE ? END,
            first_sent_at = ?, last_sent_at = ?, updated_at = ?
        WHERE id = ?
        """,
        (next_status, first_sent, sent_at, sent_at, guest_id),
    )
    record_send_log(guest_id, user["id"], sent_at)
    get_db().commit()
    return redirect(whatsapp_link(guest))


@app.route("/bulk-send", methods=["POST"])
@login_required()
def bulk_send():
    user = current_user()
    ids = [int(x) for x in request.form.getlist("guest_ids") if x.isdigit()]
    if not ids:
        flash("Lütfen en az bir davetli seçin.", "error")
        return redirect(url_for("guests"))
    placeholders = ",".join("?" for _ in ids)
    rows = get_db().execute(
        f"SELECT * FROM guests WHERE user_id = ? AND id IN ({placeholders})",
        [user["id"], *ids],
    ).fetchall()
    for guest in rows:
        sent_at = now_iso()
        first_sent = guest["first_sent_at"] or sent_at
        get_db().execute(
            """
            UPDATE guests
            SET status = CASE WHEN status IN ('attending', 'not_attending') THEN status ELSE 'sent_waiting' END,
                first_sent_at = ?, last_sent_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (first_sent, sent_at, sent_at, guest["id"]),
        )
        record_send_log(guest["id"], user["id"], sent_at)
    get_db().commit()
    links = [whatsapp_link(g) for g in rows]
    return render_template("bulk_send.html", links=links, guests=rows)


@app.route("/bulk-delete", methods=["POST"])
@login_required()
def bulk_delete():
    user = current_user()
    ids = [int(x) for x in request.form.getlist("guest_ids") if x.isdigit()]
    if not ids:
        flash("Lütfen listeden çıkarılacak davetlileri seçin.", "error")
        return redirect(url_for("guests"))
    placeholders = ",".join("?" for _ in ids)
    get_db().execute(
        f"DELETE FROM guests WHERE user_id = ? AND id IN ({placeholders})",
        [user["id"], *ids],
    )
    get_db().execute(
        f"DELETE FROM guest_send_logs WHERE user_id = ? AND guest_id IN ({placeholders})",
        [user["id"], *ids],
    )
    get_db().commit()
    flash(f"Seçili {len(ids)} davetli listeden çıkarıldı.", "success")
    return redirect(url_for("guests"))


@app.route("/attending/cancel-selected", methods=["POST"])
@login_required()
def cancel_attending_selected():
    user = current_user()
    ids = [int(x) for x in request.form.getlist("guest_ids") if x.isdigit()]
    if not ids:
        flash("Lütfen listeden çıkarılacak katılımcıları seçin.", "error")
        return redirect(url_for("attending"))
    placeholders = ",".join("?" for _ in ids)
    get_db().execute(
        f"""
        UPDATE guests
        SET status = 'cancelled_after_attending', updated_at = ?
        WHERE user_id = ? AND status = 'attending' AND id IN ({placeholders})
        """,
        [now_iso(), user["id"], *ids],
    )
    get_db().commit()
    flash(f"Seçili {len(ids)} kişi katılımcılar listesinden çıkarıldı.", "success")
    return redirect(url_for("attending"))


@app.route("/guests/<int:guest_id>/manual-attending", methods=["POST"])
@login_required()
def manual_attending(guest_id):
    user = current_user()
    guest = get_db().execute(
        "SELECT * FROM guests WHERE id = ? AND user_id = ?", (guest_id, user["id"])
    ).fetchone()
    if not guest:
        flash("Davetli bulunamadı.", "error")
        return redirect(url_for("home"))
    get_db().execute(
        "UPDATE guests SET status = 'attending', responded_at = ?, updated_at = ? WHERE id = ?",
        (now_iso(), now_iso(), guest_id),
    )
    get_db().commit()
    flash("Davetli katılımı manuel onaylandı.", "success")
    return redirect(request.form.get("next") or url_for("attending"))


@app.route("/guests/<int:guest_id>/attendance-edit")
@login_required()
def attendance_edit(guest_id):
    user = current_user()
    guest = get_db().execute(
        "SELECT * FROM guests WHERE id = ? AND user_id = ?", (guest_id, user["id"])
    ).fetchone()
    if not guest:
        flash("Davetli bulunamadı.", "error")
        return redirect(url_for("home"))
    next_url = request.args.get("next") or url_for("home")
    return render_template("attendance_edit.html", guest=guest, next_url=next_url)


@app.route("/excel-import", methods=["POST"])
@login_required()
def excel_import():
    user = current_user()
    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("Lütfen Excel dosyası seçin.", "error")
        return redirect(url_for("guests"))
    if Path(file.filename).suffix.lower() != ".xlsx":
        flash("Lütfen XLSX formatında Excel dosyası yükleyin.", "error")
        return redirect(url_for("guests"))
    added = skipped = 0
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 2 or not row[0] or not row[1]:
            continue
        ok, _ = add_guest(user["id"], row[0], row[1])
        added += 1 if ok else 0
        skipped += 0 if ok else 1
    get_db().commit()
    flash(f"Excel yükleme tamamlandı. Eklenen: {added}, atlanan: {skipped}", "success")
    return redirect(url_for("guests"))


@app.route("/attending")
@login_required()
def attending():
    user = current_user()
    return render_template(
        "list.html",
        title="Katılımcılar",
        guests=load_guests(user["id"], "attending"),
        mode="attending",
    )


@app.route("/attending/add", methods=["POST"])
@login_required()
def add_attending():
    user = current_user()
    ok, message = add_direct_participant(
        user["id"],
        request.form.get("full_name"),
        request.form.get("phone"),
    )
    flash(message, "success" if ok else "error")
    get_db().commit()
    return redirect(url_for("attending"))


@app.route("/not-attending")
@login_required()
def not_attending():
    user = current_user()
    return render_template(
        "list.html",
        title="Katılamayanlar",
        guests=load_guests(user["id"], "not_attending"),
        mode="not_attending",
    )


@app.route("/no-response")
@login_required()
def no_response():
    user = current_user()
    return render_template(
        "list.html",
        title="Cevap Vermeyenler",
        guests=load_guests(user["id"], "no_response"),
        mode="no_response",
    )


@app.route("/phone-followups")
@login_required()
def phone_followups():
    user = current_user()
    return render_template(
        "list.html",
        title="Telefonla Aranması Gerekenler",
        guests=load_guests(user["id"], "phone_followup"),
        mode="phone_followup",
    )


@app.route("/settings", methods=["GET", "POST"])
@login_required()
def settings():
    user = current_user()
    if request.method == "POST":
        try:
            hours = max(1, int(request.form.get("hours", 48)))
        except ValueError:
            hours = 48
        get_db().execute(
            """
            INSERT INTO settings (user_id, no_response_after_hours)
            VALUES (?, ?)
            ON CONFLICT(user_id) DO UPDATE SET no_response_after_hours = excluded.no_response_after_hours
            """,
            (user["id"], hours),
        )
        get_db().commit()
        flash("Ayarlar kaydedildi.", "success")
        return redirect(url_for("settings"))
    return render_template("settings.html", hours=get_setting(user["id"]))


@app.route("/import", methods=["GET", "POST"])
@login_required()
def import_guests():
    user = current_user()
    if request.method == "POST":
        file = request.files.get("guest_file")
        if not file or not file.filename:
            flash("Lütfen dosya seçin.", "error")
            return redirect(url_for("import_guests"))
        ext = Path(file.filename).suffix.lower()
        if ext not in ALLOWED_IMPORT_EXTENSIONS:
            flash("Sadece JSON dosyası yükleyin.", "error")
            return redirect(url_for("import_guests"))
        added = skipped = 0
        data = json.load(file)
        for item in data.get("guests", data if isinstance(data, list) else []):
            ok, _ = add_guest(user["id"], item.get("full_name"), item.get("phone"))
            added += 1 if ok else 0
            skipped += 0 if ok else 1
        get_db().commit()
        flash(f"İçe aktarma tamamlandı. Eklenen: {added}, atlanan: {skipped}", "success")
        return redirect(url_for("guests"))
    return render_template("import.html")


@app.route("/export")
@login_required()
def export_guests():
    user = current_user()
    return render_template("export_confirm.html", guests=load_guests(user["id"]))


@app.route("/export/download", methods=["POST"])
@login_required()
def export_download():
    user = current_user()
    rows = load_guests(user["id"])
    payload = [
        {"full_name": r["full_name"], "phone": r["phone"], "status": display_status(r)}
        for r in rows
    ]
    path = UPLOAD_DIR / f"lcv_guests_export_{user['id']}.json"
    path.write_text(json.dumps({"guests": payload}, ensure_ascii=False, indent=2), encoding="utf-8")
    return send_from_directory(UPLOAD_DIR, path.name, as_attachment=True)


def final_participant_rows(user_id):
    return load_guests(user_id, "attending")


def build_final_json(user_id):
    rows = final_participant_rows(user_id)
    payload = {
        "final_participants": [
            {
                "full_name": row["full_name"],
                "phone": row["phone"],
                "responded_at": row["responded_at"],
            }
            for row in rows
        ]
    }
    path = UPLOAD_DIR / f"son_durum_katilimci_listesi_{user_id}.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def build_final_excel(user_id):
    rows = final_participant_rows(user_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Son Durum Katılımcılar"
    ws.append(["İsim Soyisim", "Telefon", "Katılım teyit zamanı"])
    for row in rows:
        ws.append([row["full_name"], row["phone"], row["responded_at"]])
    path = UPLOAD_DIR / f"son_durum_katilimci_listesi_{user_id}.xlsx"
    wb.save(path)
    return path


@app.route("/api/final-participants/excel")
@api_login_required
def api_final_excel():
    path = build_final_excel(g.api_user["id"])
    return send_from_directory(UPLOAD_DIR, path.name, as_attachment=True)


@app.route("/api/final-participants/json")
@api_login_required
def api_final_json():
    path = build_final_json(g.api_user["id"])
    return send_from_directory(UPLOAD_DIR, path.name, as_attachment=True)


@app.route("/share-final")
@login_required()
def share_final():
    user = current_user()
    return render_template("share_final.html", count=len(final_participant_rows(user["id"])))


@app.route("/share-final/excel")
@login_required()
def share_final_excel():
    user = current_user()
    path = build_final_excel(user["id"])
    return send_from_directory(UPLOAD_DIR, path.name, as_attachment=True)


@app.route("/share-final/json")
@login_required()
def share_final_json():
    user = current_user()
    path = build_final_json(user["id"])
    return send_from_directory(UPLOAD_DIR, path.name, as_attachment=True)


@app.route("/r/<token>", methods=["GET", "POST"])
def respond(token):
    guest = get_db().execute("SELECT * FROM guests WHERE response_token = ?", (token,)).fetchone()
    if not guest:
        return render_template("public_message.html", message="Davet linki bulunamadı.")
    inv = invitation_for(guest["user_id"])
    if request.method == "POST":
        answer = request.form.get("answer")
        status = "attending" if answer == "attending" else "not_attending"
        get_db().execute(
            "UPDATE guests SET status = ?, responded_at = ?, updated_at = ? WHERE id = ?",
            (status, now_iso(), now_iso(), guest["id"]),
        )
        get_db().commit()
        return render_template("public_message.html", message="Cevabınız alınmıştır. Teşekkür ederiz.")
    return render_template("respond.html", guest=guest, invitation_url=invitation_url(inv))


@app.route("/account-deletion", methods=["GET", "POST"])
def account_deletion():
    submitted = False
    email = ""
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        if not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email):
            flash("Lütfen geçerli bir e-posta adresi girin.", "error")
        else:
            db = get_db()
            db.execute(
                """
                INSERT INTO account_deletion_requests (email, requested_at)
                VALUES (?, ?)
                ON CONFLICT(email) DO UPDATE SET requested_at = excluded.requested_at
                """,
                (email, now_iso()),
            )
            db.commit()
            submitted = True
    return render_template("account_deletion.html", submitted=submitted, email=email)


@app.route("/privacy-policy")
def privacy_policy():
    return render_template("privacy_policy.html")


@app.route("/admin")
@login_required("super_admin")
def admin():
    db = get_db()
    users = db.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()
    deletion_requests = db.execute(
        "SELECT * FROM account_deletion_requests ORDER BY requested_at DESC, id DESC"
    ).fetchall()
    active_filter = request.args.get("type", "users")
    cards = []
    for user in users:
        if user["role"] == "super_admin":
            continue
        if active_filter == "organizers" and user["role"] != "organizer":
            continue
        if active_filter == "users" and user["role"] != "couple":
            continue
        cards.append({"user": user, "counts": guest_counts(user["id"])})
    return render_template(
        "admin.html",
        users=cards,
        active_filter=active_filter,
        deletion_requests=deletion_requests,
    )


@app.route("/admin/account-deletion-requests/<int:request_id>/complete", methods=["POST"])
@login_required("super_admin")
def admin_complete_account_deletion_request(request_id):
    db = get_db()
    deletion_request = db.execute(
        "SELECT * FROM account_deletion_requests WHERE id = ?", (request_id,)
    ).fetchone()
    if not deletion_request:
        flash("Hesap silme talebi bulunamadı.", "error")
        return redirect(url_for("admin"))
    db.execute("DELETE FROM account_deletion_requests WHERE id = ?", (request_id,))
    db.commit()
    flash("Tamamlanan hesap silme talebi kayıtlardan kaldırıldı.", "success")
    return redirect(url_for("admin"))


@app.route("/admin/accounts/create", methods=["POST"])
@login_required("super_admin")
def admin_create_account():
    role = request.form.get("role", "couple")
    try:
        _user_id, password = create_managed_account(request.form.get("email"), role)
    except (ValueError, LookupError) as exc:
        flash(str(exc), "error")
        return redirect(url_for("admin", type="organizers" if role == "organizer" else "users"))
    label = "Organizatör" if role == "organizer" else "Normal kullanıcı"
    flash(f"{label} oluşturuldu. Şifre: {password}", "success")
    return redirect(url_for("admin", type="organizers" if role == "organizer" else "users"))


@app.route("/admin/organizers/<int:user_id>")
@login_required("super_admin")
def admin_organizer(user_id):
    organizer = get_db().execute(
        "SELECT * FROM users WHERE id = ? AND role = 'organizer'", (user_id,)
    ).fetchone()
    if not organizer:
        flash("Organizatör bulunamadı.", "error")
        return redirect(url_for("admin", type="organizers"))
    users = get_db().execute(
        "SELECT * FROM users WHERE role = 'couple' AND managed_by_user_id = ? ORDER BY created_at DESC, id DESC",
        (user_id,),
    ).fetchall()
    cards = [{"user": user, "counts": guest_counts(user["id"])} for user in users]
    return render_template("organizer.html", organizer=organizer, users=cards, super_view=True)


@app.route("/organizer")
@login_required("organizer")
def organizer_dashboard():
    organizer = current_user()
    users = get_db().execute(
        "SELECT * FROM users WHERE role = 'couple' AND managed_by_user_id = ? ORDER BY created_at DESC, id DESC",
        (organizer["id"],),
    ).fetchall()
    cards = [{"user": user, "counts": guest_counts(user["id"])} for user in users]
    return render_template("organizer.html", organizer=organizer, users=cards, super_view=False)


@app.route("/organizer/users/create", methods=["POST"])
@login_required("organizer")
def organizer_create_user():
    organizer = current_user()
    try:
        _user_id, password = create_managed_account(
            request.form.get("email"), "couple", organizer["id"]
        )
    except (ValueError, LookupError) as exc:
        flash(str(exc), "error")
        return redirect(url_for("organizer_dashboard"))
    flash(f"Kullanıcı oluşturuldu. Şifre: {password}", "success")
    return redirect(url_for("organizer_dashboard"))


@app.route("/users/<int:user_id>/readonly")
@login_required()
def user_readonly(user_id):
    viewer = current_user()
    user = get_db().execute(
        "SELECT * FROM users WHERE id = ? AND role = 'couple'", (user_id,)
    ).fetchone()
    allowed = user and (
        viewer["role"] == "super_admin"
        or (viewer["role"] == "organizer" and user["managed_by_user_id"] == viewer["id"])
    )
    if not allowed:
        flash("Bu kullanıcıyı görüntüleme yetkiniz yok.", "error")
        return redirect(url_for("home"))
    context = owner_context(user_id)
    back_url = (
        url_for("admin_organizer", user_id=user["managed_by_user_id"])
        if viewer["role"] == "super_admin" and user["managed_by_user_id"]
        else url_for("admin", type="users") if viewer["role"] == "super_admin"
        else url_for("organizer_dashboard")
    )
    return render_template(
        "readonly_user.html",
        **context,
        back_url=back_url,
        phone_followups=load_guests(user_id, "phone_followup"),
        final_participants=final_participant_rows(user_id),
        invitation_message=get_invitation_message(user_id),
        no_response_after_hours=get_setting(user_id),
    )


@app.route("/admin/beta-users/create", methods=["POST"])
@login_required("super_admin")
def admin_create_beta_user():
    try:
        _user_id, password = create_beta_user(request.form.get("email"))
    except (ValueError, LookupError) as exc:
        flash(str(exc), "error")
        return redirect(url_for("admin", type="beta"))

    flash(f"Beta kullanıcı oluşturuldu. Şifre: {password}", "success")
    return redirect(url_for("admin", type="beta"))


@app.route("/admin/customers/<int:user_id>")
@login_required("super_admin")
def admin_customer(user_id):
    context = owner_context(user_id)
    if not context:
        flash("Müşteri bulunamadı.", "error")
        return redirect(url_for("admin"))
    return render_template("admin_customer.html", **context)


@app.route("/admin/customers/<int:user_id>/access", methods=["POST"])
@login_required("super_admin")
def admin_customer_access(user_id):
    user = get_db().execute(
        "SELECT * FROM users WHERE id = ? AND role IN ('couple', 'organizer')", (user_id,)
    ).fetchone()
    if not user:
        flash("Müşteri bulunamadı.", "error")
        return redirect(url_for("admin"))
    action = request.form.get("action")
    if action == "beta_on":
        get_db().execute(
            "UPDATE users SET is_beta_user = 1, is_access_blocked = 0 WHERE id = ?",
            (user_id,),
        )
        message = "Kullanıcı beta kullanıcı yapıldı."
    elif action == "beta_off":
        get_db().execute(
            "UPDATE users SET is_beta_user = 0, is_access_blocked = 0 WHERE id = ?",
            (user_id,),
        )
        message = "Beta kullanıcı durumu kaldırıldı."
    elif action == "access_block":
        get_db().execute(
            "UPDATE users SET is_access_blocked = 1 WHERE id = ?",
            (user_id,),
        )
        message = "Beta kullanıcının erişimi durduruldu."
    elif action == "access_unblock":
        get_db().execute(
            "UPDATE users SET is_access_blocked = 0 WHERE id = ?",
            (user_id,),
        )
        message = "Beta kullanıcının erişimi tekrar aktive edildi."
    else:
        flash("Geçersiz erişim işlemi.", "error")
        return redirect(url_for("admin", type="organizers" if user["role"] == "organizer" else "users"))
    get_db().commit()
    flash(message, "success")
    return redirect(url_for("admin", type="organizers" if user["role"] == "organizer" else "users"))


@app.route("/admin/accounts/<int:user_id>/delete", methods=["POST"])
@login_required("super_admin")
def admin_delete_account(user_id):
    user = delete_managed_account_records(user_id)
    if not user:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin"))
    flash("Kullanıcı ve bağlı kayıtları kalıcı olarak silindi.", "success")
    return redirect(url_for("admin", type="organizers" if user["role"] == "organizer" else "users"))


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


@app.context_processor
def inject_helpers():
    return {
        "display_status": display_status,
        "is_no_response": is_no_response,
        "is_phone_call_needed": is_phone_call_needed,
        "personal_link": personal_link,
        "whatsapp_link": whatsapp_link,
        "send_history": send_history,
        "user_access_payload": user_access_payload,
    }


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5062")), debug=True)
